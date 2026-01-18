import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import sympy

EQUATION_STR = "x^4 + 2*x^2 - x - 3"
INPUT_X0 = 1
TOLERANCE = 0.000001
DECIMALS = 8

print(f"สมการตั้งต้น f(x): {EQUATION_STR}")
try:
    x_sym = sympy.symbols('x')
    
    eq_for_sympy = EQUATION_STR.replace('^', '**')
    
    expr = sympy.sympify(eq_for_sympy)
    
    diff_expr = sympy.diff(expr, x_sym)
    
    DERIVATIVE_STR = str(diff_expr).replace('**', '^')
    print(f"คำนวณ f'(x) อัตโนมัติได้: {DERIVATIVE_STR}")
    
except Exception as e:
    print(f"เกิดข้อผิดพลาดในการ Diff: {e}")
    DERIVATIVE_STR = "4*x^3 + 4*x - 1"

OUTPUT_FILENAME = "Newton_Result.xlsx"

def format_excel_formula(equation, cell_ref):
    formula = equation.lower()
    formula = formula.replace('sin', 'SIN')
    formula = formula.replace('cos', 'COS')
    formula = formula.replace('tan', 'TAN')
    formula = formula.replace('sqrt', 'SQRT')
    formula = formula.replace('abs', 'ABS')
    formula = formula.replace('log', 'LOG')
    formula = formula.replace('exp', 'EXP')
    
    formula = re.sub(r'(?<![a-zA-Z])x(?![a-zA-Z])', cell_ref, formula)
    
    return "=" + formula

def evaluate_func(eq, x):
    eq_py = eq.replace('^', '**')
    for func in ['sin', 'cos', 'tan', 'sqrt', 'exp', 'log', 'abs']:
        eq_py = eq_py.replace(func, f'math.{func}')
    try:
        return eval(eq_py)
    except:
        return 0

def generate_excel():
    print(f"กำลังคำนวณและสร้างไฟล์ {OUTPUT_FILENAME}...")
    
    x_curr = INPUT_X0
    total_loops = 0
    
    for i in range(1, 100):
        fx = evaluate_func(EQUATION_STR, x_curr)
        fpx = evaluate_func(DERIVATIVE_STR, x_curr)
        
        if fpx == 0: 
            print("Warning: Derivative is zero. Stopping.")
            break
            
        x_next = x_curr - (fx / fpx)
        
        error = abs(x_next - x_curr)
        
        x_curr = x_next
        total_loops = i
        
        if error < TOLERANCE:
            break

    print(f"คำนวณจบที่ {total_loops} รอบ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Newton Method"

    ROW_HEADER = 1
    ROW_X_OLD = 2
    ROW_FX = 3
    ROW_FPX = 4
    ROW_X_NEW = 5
    ROW_ERR = 6
    ROW_CHECK = 7

    labels = {
        ROW_X_OLD: "x_(i-1)", 
        ROW_FX: "f(x_(i-1))", 
        ROW_FPX: "f(x_(i-1))'", 
        ROW_X_NEW: "x_(i+1)", 
        ROW_ERR: "error", 
        ROW_CHECK: "check round"
    }
    
    ws.cell(row=ROW_HEADER, column=1, value="Variables")
    for r, text in labels.items():
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=True, color="3730A3")
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    ws.column_dimensions['B'].width = 3

    start_col = 3
    number_format = '0.' + '0' * DECIMALS

    for i in range(total_loops):
        col_idx = start_col + i
        col_letter = get_column_letter(col_idx)
        prev_col_letter = get_column_letter(col_idx - 1)
        
        header_cell = ws.cell(row=ROW_HEADER, column=col_idx, value=f"x{i+1}")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        if i == 0:
            ws.cell(row=ROW_X_OLD, column=col_idx, value=INPUT_X0).number_format = number_format
        else:
            ws.cell(row=ROW_X_OLD, column=col_idx, value=f'={prev_col_letter}{ROW_X_NEW}').number_format = number_format

        ws.cell(row=ROW_FX, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_X_OLD}")).number_format = number_format

        ws.cell(row=ROW_FPX, column=col_idx, value=format_excel_formula(DERIVATIVE_STR, f"{col_letter}{ROW_X_OLD}")).number_format = number_format

        ws.cell(row=ROW_X_NEW, column=col_idx, value=f'={col_letter}{ROW_X_OLD} - {col_letter}{ROW_FX}/{col_letter}{ROW_FPX}').number_format = number_format

        ws.cell(row=ROW_ERR, column=col_idx, value=f'=ABS({col_letter}{ROW_X_NEW} - {col_letter}{ROW_X_OLD})').number_format = number_format

        meta_col_letter = get_column_letter(start_col + total_loops + 3)
        tol_cell_ref = f"${meta_col_letter}$6"

        check_cell = ws.cell(row=ROW_CHECK, column=col_idx, value=f'=IF({col_letter}{ROW_ERR}<{tol_cell_ref}, "TRUE", "FALSE")')
        
    meta_col = start_col + total_loops + 1
    ws.column_dimensions[get_column_letter(meta_col)].width = 2
    
    meta_col_idx = meta_col + 1
    ws.column_dimensions[get_column_letter(meta_col_idx)].width = 20
    ws.column_dimensions[get_column_letter(meta_col_idx+1)].width = 20

    for r in range(1, 7):
        ws.cell(row=r, column=meta_col_idx).font = Font(bold=True)

    ws.cell(row=1, column=meta_col_idx, value="Parameters")
    ws.cell(row=1, column=meta_col_idx+1, value="Values").font = Font(bold=True)
    
    ws.cell(row=2, column=meta_col_idx, value="f(x)")
    ws.cell(row=2, column=meta_col_idx+1, value=EQUATION_STR)
    
    ws.cell(row=3, column=meta_col_idx, value="f'(x)")
    ws.cell(row=3, column=meta_col_idx+1, value=DERIVATIVE_STR)
    
    ws.cell(row=4, column=meta_col_idx, value="x0")
    ws.cell(row=4, column=meta_col_idx+1, value=INPUT_X0)
    
    ws.cell(row=6, column=meta_col_idx, value="Tolerance")
    ws.cell(row=6, column=meta_col_idx+1, value=TOLERANCE)

    ws.column_dimensions['A'].width = 15
    for i in range(total_loops):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 16

    wb.save(OUTPUT_FILENAME)
    print(f"เสร็จสิ้น! บันทึกไฟล์ที่: {OUTPUT_FILENAME}")
    print("เปิดไฟล์ Excel แล้วสูตรจะทำงานทันทีครับ")

if __name__ == "__main__":
    generate_excel()