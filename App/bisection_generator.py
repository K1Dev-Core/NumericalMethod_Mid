import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

EQUATION_STR = "x^3 + 3*x^2 - 1"
INPUT_A = 0
INPUT_B = 1
TOLERANCE = 0.000001
DECIMALS = 10

OUTPUT_FILENAME = "Bisection_Result_Python.xlsx"

def format_excel_formula(equation, cell_ref):
    formula = equation.lower()
    formula = formula.replace('sin', 'SIN')
    formula = formula.replace('cos', 'COS')
    formula = formula.replace('tan', 'TAN')
    formula = formula.replace('sqrt', 'SQRT')
    formula = formula.replace('abs', 'ABS')
    formula = formula.replace('log', 'LOG')
    formula = formula.replace('exp', 'EXP')
    
    import re
    formula = re.sub(r'(?<![a-zA-Z])x(?![a-zA-Z])', cell_ref, formula)
    
    return "=" + formula

def evaluate_func(eq, x):
    eq_py = eq.replace('^', '**')
    for func in ['sin', 'cos', 'tan', 'sqrt', 'exp', 'log', 'abs']:
        eq_py = eq_py.replace(func, f'math.{func}')
    try:
        return eval(eq_py)
    except:
        return 0

def generate_excel():
    print(f"กำลังคำนวณและสร้างไฟล์ {OUTPUT_FILENAME}...")
    
    a = INPUT_A
    b = INPUT_B
    iteration_data = []
    
    for i in range(1, 100):
        m = (a + b) / 2
        fa = evaluate_func(EQUATION_STR, a)
        fm = evaluate_func(EQUATION_STR, m)
        fafa = fm * fa
        error = abs(b - a) / 2
        
        iteration_data.append(i)
        
        if error < TOLERANCE:
            break
            
        if fafa < 0:
            b = m
        else:
            a = m

    total_loops = len(iteration_data)
    print(f"คำนวณจบที่ {total_loops} รอบ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bisection Method"

    ROW_HEADER = 1
    ROW_A = 2
    ROW_B = 3
    ROW_FA = 4
    ROW_FB = 5
    ROW_M = 6
    ROW_FM = 7
    ROW_FAFA = 8
    ROW_ERR = 9
    ROW_CHECK = 10
    ROW_RANGE = 11
    ROW_NEXT = 12

    labels = {
        ROW_A: "a", ROW_B: "b", ROW_FA: "f(a)", ROW_FB: "f(b)",
        ROW_M: "m", ROW_FM: "f(m)", ROW_FAFA: "f(m)f(a)", 
        ROW_ERR: "err", ROW_CHECK: "check err", 
        ROW_RANGE: "อยู่ในช่วง?", ROW_NEXT: "ต่อไปอะไรเปลี่ยน"
    }
    
    ws.cell(row=ROW_HEADER, column=1, value="Variables")
    for r, text in labels.items():
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=True, color="3730A3")
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    ws.column_dimensions['B'].width = 3

    start_col = 3
    
    number_format = '0.' + '0' * DECIMALS

    for i in range(total_loops):
        col_idx = start_col + i
        col_letter = get_column_letter(col_idx)
        prev_col_letter = get_column_letter(col_idx - 1)
        
        header_cell = ws.cell(row=ROW_HEADER, column=col_idx, value=f"i = {i+1}")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        if i == 0:
            ws.cell(row=ROW_A, column=col_idx, value=INPUT_A).number_format = number_format
            ws.cell(row=ROW_B, column=col_idx, value=INPUT_B).number_format = number_format
        else:
            ws.cell(row=ROW_A, column=col_idx, value=f'=IF({prev_col_letter}{ROW_FAFA}>0, {prev_col_letter}{ROW_M}, {prev_col_letter}{ROW_A})').number_format = number_format
            ws.cell(row=ROW_B, column=col_idx, value=f'=IF({prev_col_letter}{ROW_FAFA}<0, {prev_col_letter}{ROW_M}, {prev_col_letter}{ROW_B})').number_format = number_format

        ws.cell(row=ROW_FA, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_A}")).number_format = number_format
        ws.cell(row=ROW_FB, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_B}")).number_format = number_format

        ws.cell(row=ROW_M, column=col_idx, value=f'=({col_letter}{ROW_A}+{col_letter}{ROW_B})/2').number_format = number_format
        
        ws.cell(row=ROW_FM, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_M}")).number_format = number_format

        ws.cell(row=ROW_FAFA, column=col_idx, value=f'={col_letter}{ROW_FM}*{col_letter}{ROW_FA}').number_format = number_format

        ws.cell(row=ROW_ERR, column=col_idx, value=f'=ABS({col_letter}{ROW_B}-{col_letter}{ROW_A})/2').number_format = number_format

        meta_col_letter = get_column_letter(start_col + total_loops + 3)
        tol_cell_ref = f"${meta_col_letter}$5"

        check_cell = ws.cell(row=ROW_CHECK, column=col_idx, value=f'=IF({col_letter}{ROW_ERR}<{tol_cell_ref}, "TRUE", "FALSE")')
        
        ws.cell(row=ROW_RANGE, column=col_idx, value=f'=IF({col_letter}{ROW_FAFA}<0, "อยู่ในช่วง", "อยู่นอกช่วง")')

        next_cell = ws.cell(row=ROW_NEXT, column=col_idx, value=f'=IF({col_letter}{ROW_ERR}<{tol_cell_ref}, "จบ", IF({col_letter}{ROW_FAFA}<0, "B", "A"))')
        next_cell.font = Font(bold=True, color="4F46E5")

    meta_col = start_col + total_loops + 1
    ws.column_dimensions[get_column_letter(meta_col)].width = 2
    
    meta_col_idx = meta_col + 1
    ws.column_dimensions[get_column_letter(meta_col_idx)].width = 20
    ws.column_dimensions[get_column_letter(meta_col_idx+1)].width = 20

    ws.cell(row=1, column=meta_col_idx, value="Parameters").font = Font(bold=True)
    ws.cell(row=1, column=meta_col_idx+1, value="Values").font = Font(bold=True)

    ws.cell(row=2, column=meta_col_idx, value="Equation")
    ws.cell(row=2, column=meta_col_idx+1, value=EQUATION_STR)

    ws.cell(row=3, column=meta_col_idx, value="Initial a")
    ws.cell(row=3, column=meta_col_idx+1, value=INPUT_A)

    ws.cell(row=4, column=meta_col_idx, value="Initial b")
    ws.cell(row=4, column=meta_col_idx+1, value=INPUT_B)

    ws.cell(row=5, column=meta_col_idx, value="Tolerance")
    ws.cell(row=5, column=meta_col_idx+1, value=TOLERANCE)

    ws.column_dimensions['A'].width = 15
    for i in range(total_loops):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 14

    wb.save(OUTPUT_FILENAME)
    print(f"เสร็จสิ้น! บันทึกไฟล์ที่: {OUTPUT_FILENAME}")
    print("เปิดไฟล์ Excel แล้วสูตรจะทำงานทันทีครับ")

if __name__ == "__main__":
    generate_excel()