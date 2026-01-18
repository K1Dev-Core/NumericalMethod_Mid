import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re

EQUATION_STR = "sin(x) - x^2"
INPUT_X0 = 0.75
INPUT_X1 = 1.0
TOLERANCE = 0.001 
DECIMALS = 6

OUTPUT_FILENAME = "Secant_Result.xlsx"

def format_excel_formula(equation, cell_ref):
    formula = equation.lower()
    formula = formula.replace('sin', 'SIN')
    formula = formula.replace('cos', 'COS')
    formula = formula.replace('tan', 'TAN')
    formula = formula.replace('sqrt', 'SQRT')
    formula = formula.replace('abs', 'ABS')
    formula = formula.replace('log', 'LOG')
    formula = formula.replace('exp', 'EXP')
    
    formula = re.sub(r'(?<![a-zA-Z])x(?![a-zA-Z])', cell_ref, formula)
    return "=" + formula

def evaluate_func(eq, x):
    eq_py = eq.replace('^', '**')
    for func in ['sin', 'cos', 'tan', 'sqrt', 'exp', 'log', 'abs']:
        eq_py = eq_py.replace(func, f'math.{func}')
    try:
        return eval(eq_py)
    except:
        return 0

def generate_excel():
    print(f"กำลังคำนวณและสร้างไฟล์ {OUTPUT_FILENAME}...")
    
    x_prev = INPUT_X0
    x_curr = INPUT_X1
    total_loops = 0
    
    for i in range(1, 100):
        fx_curr = evaluate_func(EQUATION_STR, x_curr)
        fx_prev = evaluate_func(EQUATION_STR, x_prev)
        
        if (fx_curr - fx_prev) == 0:
            print("Warning: Division by zero (f(xi) = f(xi-1)). Stopping.")
            break
            
        x_next = x_curr - (fx_curr * (x_curr - x_prev)) / (fx_curr - fx_prev)
        
        error = abs(x_next - x_curr)
        
        x_prev = x_curr
        x_curr = x_next
        total_loops = i
        
        if error < TOLERANCE:
            break

    print(f"คำนวณจบที่ {total_loops} รอบ")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Secant Method"

    ROW_HEADER = 1
    ROW_XI = 2
    ROW_XI_PREV = 3
    ROW_FXI = 4
    ROW_FXI_PREV = 5
    ROW_DIFF_X = 6
    ROW_XI_NEXT = 7
    ROW_ERR = 8
    ROW_CHECK = 9

    labels = {
        ROW_XI: "x_i",
        ROW_XI_PREV: "x_(i-1)",
        ROW_FXI: "f(x_i)",
        ROW_FXI_PREV: "f(x_(i-1))",
        ROW_DIFF_X: "(x_i) - x_(i-1)",
        ROW_XI_NEXT: "x_i+1",
        ROW_ERR: "error",
        ROW_CHECK: "check round"
    }
    
    ws.cell(row=ROW_HEADER, column=1, value="Variables")
    for r, text in labels.items():
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=True, color="3730A3")
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    ws.column_dimensions['B'].width = 3

    start_col = 3
    number_format = '0.' + '0' * DECIMALS

    for i in range(total_loops):
        col_idx = start_col + i
        col_letter = get_column_letter(col_idx)
        prev_col_letter = get_column_letter(col_idx - 1)
        
        header_cell = ws.cell(row=ROW_HEADER, column=col_idx, value=f"x{i+2}")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        if i == 0:
            ws.cell(row=ROW_XI, column=col_idx, value=INPUT_X1).number_format = number_format
        else:
            ws.cell(row=ROW_XI, column=col_idx, value=f'={prev_col_letter}{ROW_XI_NEXT}').number_format = number_format

        if i == 0:
            ws.cell(row=ROW_XI_PREV, column=col_idx, value=INPUT_X0).number_format = number_format
        else:
            ws.cell(row=ROW_XI_PREV, column=col_idx, value=f'={prev_col_letter}{ROW_XI}').number_format = number_format

        ws.cell(row=ROW_FXI, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_XI}")).number_format = number_format

        ws.cell(row=ROW_FXI_PREV, column=col_idx, value=format_excel_formula(EQUATION_STR, f"{col_letter}{ROW_XI_PREV}")).number_format = number_format

        ws.cell(row=ROW_DIFF_X, column=col_idx, value=f'={col_letter}{ROW_XI}-{col_letter}{ROW_XI_PREV}').number_format = number_format

        secant_formula = f'={col_letter}{ROW_XI} - {col_letter}{ROW_FXI} * {col_letter}{ROW_DIFF_X} / ({col_letter}{ROW_FXI} - {col_letter}{ROW_FXI_PREV})'
        ws.cell(row=ROW_XI_NEXT, column=col_idx, value=secant_formula).number_format = number_format

        ws.cell(row=ROW_ERR, column=col_idx, value=f'=ABS({col_letter}{ROW_XI_NEXT}-{col_letter}{ROW_XI})').number_format = number_format

        meta_col_letter = get_column_letter(start_col + total_loops + 3)
        tol_cell_ref = f"${meta_col_letter}$5"

        ws.cell(row=ROW_CHECK, column=col_idx, value=f'=IF({col_letter}{ROW_ERR}<{tol_cell_ref}, "TRUE", "FALSE")')

    meta_col = start_col + total_loops + 1
    ws.column_dimensions[get_column_letter(meta_col)].width = 2
    
    meta_col_idx = meta_col + 1
    ws.column_dimensions[get_column_letter(meta_col_idx)].width = 20
    ws.column_dimensions[get_column_letter(meta_col_idx+1)].width = 20

    for r in range(1, 7):
        ws.cell(row=r, column=meta_col_idx).font = Font(bold=True)

    ws.cell(row=1, column=meta_col_idx, value="Parameters")
    ws.cell(row=1, column=meta_col_idx+1, value="Values").font = Font(bold=True)
    
    ws.cell(row=2, column=meta_col_idx, value="f(x)")
    ws.cell(row=2, column=meta_col_idx+1, value=EQUATION_STR)
    
    ws.cell(row=3, column=meta_col_idx, value="x0")
    ws.cell(row=3, column=meta_col_idx+1, value=INPUT_X0)
    
    ws.cell(row=4, column=meta_col_idx, value="x1")
    ws.cell(row=4, column=meta_col_idx+1, value=INPUT_X1)
    
    ws.cell(row=5, column=meta_col_idx, value="Tolerance")
    ws.cell(row=5, column=meta_col_idx+1, value=TOLERANCE)

    ws.column_dimensions['A'].width = 18
    for i in range(total_loops):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 16

    wb.save(OUTPUT_FILENAME)
    print(f"เสร็จสิ้น! บันทึกไฟล์ที่: {OUTPUT_FILENAME}")

if __name__ == "__main__":
    generate_excel()