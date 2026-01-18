import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MATRIX_A = [
    [1, 2, 1],
    [2, 2, 3],
    [-1, -3, 0]
]
VECTOR_B = [0, 3, 2]

OUTPUT_FILENAME = "Cramer_Rule_Result.xlsx"

def set_border(ws, row_start, col_start, row_end, col_end):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = border

def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

def style_label(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

def generate_cramer_excel():
    print(f"กำลังสร้างไฟล์ {OUTPUT_FILENAME}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cramer Rule"

    ws.cell(row=1, column=1, value="System: AX = B").font = Font(size=14, bold=True)
    
    ws.cell(row=2, column=1, value="Matrix A")
    style_header(ws.cell(row=2, column=1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    
    ws.cell(row=2, column=4, value="X")
    style_header(ws.cell(row=2, column=4))
    
    ws.cell(row=2, column=5, value="Vector b")
    style_header(ws.cell(row=2, column=5))

    for r in range(3):
        for c in range(3):
            cell = ws.cell(row=3+r, column=1+c, value=MATRIX_A[r][c])
            cell.alignment = Alignment(horizontal='center')
    
    ws.cell(row=3, column=4, value="x1")
    ws.cell(row=4, column=4, value="x2")
    ws.cell(row=5, column=4, value="x3")
    
    for r in range(3):
        cell = ws.cell(row=3+r, column=5, value=VECTOR_B[r])
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True, color="DC2626")

    set_border(ws, 2, 1, 5, 5)
    
    def create_matrix_block(start_row, title, col1_ref, col2_ref, col3_ref):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12, color="3730A3")
        
        ws.cell(row=start_row+1, column=1, value="Matrix:")
        
        ws.cell(row=start_row+2, column=1, value=f"={col1_ref}3")
        ws.cell(row=start_row+2, column=2, value=f"={col2_ref}3")
        ws.cell(row=start_row+2, column=3, value=f"={col3_ref}3")
        ws.cell(row=start_row+3, column=1, value=f"={col1_ref}4")
        ws.cell(row=start_row+3, column=2, value=f"={col2_ref}4")
        ws.cell(row=start_row+3, column=3, value=f"={col3_ref}4")
        ws.cell(row=start_row+4, column=1, value=f"={col1_ref}5")
        ws.cell(row=start_row+4, column=2, value=f"={col2_ref}5")
        ws.cell(row=start_row+4, column=3, value=f"={col3_ref}5")
        
        set_border(ws, start_row+2, 1, start_row+4, 3)
        
        det_label_row = start_row+2
        det_val_row = start_row+3
        
        ws.cell(row=det_label_row, column=5, value="Determinant").font = Font(bold=True)
        
        range_str = f"A{start_row+2}:C{start_row+4}"
        ws.cell(row=det_val_row, column=5, value=f"=MDETERM({range_str})")
        ws.cell(row=det_val_row, column=5).number_format = '0.00'
        ws.cell(row=det_val_row, column=5).font = Font(bold=True, color="0000FF")

        return f"E{det_val_row}"

    current_row = 7
    ref_det_A = create_matrix_block(current_row, "1. Find det(A)", "A", "B", "C")
    
    current_row += 6
    ref_det_x1 = create_matrix_block(current_row, "2. Find x1 -> det(x1)", "E", "B", "C")
    
    ws.cell(row=current_row+2, column=7, value="x1 = det(x1) / det(A)")
    ws.cell(row=current_row+3, column=7, value=f"={ref_det_x1}/{ref_det_A}")
    ws.cell(row=current_row+3, column=7).font = Font(bold=True, size=12, color="16A34A")
    set_border(ws, current_row+2, 7, current_row+3, 7)

    current_row += 6
    ref_det_x2 = create_matrix_block(current_row, "3. Find x2 -> det(x2)", "A", "E", "C")

    ws.cell(row=current_row+2, column=7, value="x2 = det(x2) / det(A)")
    ws.cell(row=current_row+3, column=7, value=f"={ref_det_x2}/{ref_det_A}")
    ws.cell(row=current_row+3, column=7).font = Font(bold=True, size=12, color="16A34A")
    set_border(ws, current_row+2, 7, current_row+3, 7)

    current_row += 6
    ref_det_x3 = create_matrix_block(current_row, "4. Find x3 -> det(x3)", "A", "B", "E")

    ws.cell(row=current_row+2, column=7, value="x3 = det(x3) / det(A)")
    ws.cell(row=current_row+3, column=7, value=f"={ref_det_x3}/{ref_det_A}")
    ws.cell(row=current_row+3, column=7).font = Font(bold=True, size=12, color="16A34A")
    set_border(ws, current_row+2, 7, current_row+3, 7)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['G'].width = 20

    wb.save(OUTPUT_FILENAME)
    print(f"เสร็จสิ้น! บันทึกไฟล์ที่: {OUTPUT_FILENAME}")
    print("เปิดไฟล์ Excel แล้วสูตรคำนวณ Cramer's Rule จะทำงานทันทีครับ")

if __name__ == "__main__":
    generate_cramer_excel()